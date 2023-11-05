import openpyxl
from pathlib import Path


class XlsxReader:
    def __init__(self, path: Path):
        self.workbook = openpyxl.load_workbook(path)
        self.worksheet = self.workbook.active

        self.row_size = self.worksheet.max_row
        self.column_size = self.worksheet.max_column

    def read(self, row: int) -> list:
        data = []

        for column in range(1, self.column_size + 1):
            data.append(self.worksheet.cell(row=row, column=column))

        return data
